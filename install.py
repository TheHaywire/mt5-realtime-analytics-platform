#!/usr/bin/env python3
"""
Automated Installation Script for Gold Trading Statistical Analysis
Handles dependency installation, directory setup, and system verification
"""

import sys
import subprocess
import os
from pathlib import Path
import platform

def print_header():
    """Print installation header"""
    print("=" * 60)
    print("🏆 Gold Trading Statistical Analysis - Installation")
    print("=" * 60)
    print()

def check_python_version():
    """Check if Python version is compatible"""
    print("📋 Checking Python version...")
    
    version = sys.version_info
    if version.major < 3 or (version.major == 3 and version.minor < 8):
        print("❌ Error: Python 3.8 or higher required")
        print(f"   Current version: {version.major}.{version.minor}.{version.micro}")
        print("   Please upgrade Python and try again")
        return False
    
    print(f"✅ Python {version.major}.{version.minor}.{version.micro} - Compatible")
    return True

def check_operating_system():
    """Check operating system compatibility"""
    print("💻 Checking operating system...")
    
    os_name = platform.system()
    if os_name == "Windows":
        print("✅ Windows OS detected - Full MT5 support available")
        return True
    else:
        print(f"⚠️  {os_name} OS detected - MT5 support limited")
        print("   You can still use offline analysis features")
        return True

def install_requirements():
    """Install Python packages from requirements file"""
    print("📦 Installing Python dependencies...")
    
    requirements_file = "requirements_blog.txt"
    if not os.path.exists(requirements_file):
        print(f"❌ Error: {requirements_file} not found")
        return False
    
    try:
        # Install packages
        subprocess.run([
            sys.executable, "-m", "pip", "install", "-r", requirements_file
        ], check=True, capture_output=True, text=True)
        
        print("✅ Dependencies installed successfully")
        return True
        
    except subprocess.CalledProcessError as e:
        print("❌ Error installing dependencies:")
        print(f"   {e.stderr}")
        return False

def create_directory_structure():
    """Create necessary directories"""
    print("📁 Creating directory structure...")
    
    directories = [
        "blog_data_exports",
        "blog_data_exports/raw_data", 
        "blog_data_exports/analysis_results",
        "blog_data_exports/trading_logs",
        "blog_data_exports/charts",
        "blog_data_exports/downloadable_tools",
        "config",
        "logs",
        "temp"
    ]
    
    created_count = 0
    for directory in directories:
        dir_path = Path(directory)
        if not dir_path.exists():
            dir_path.mkdir(parents=True, exist_ok=True)
            created_count += 1
    
    print(f"✅ Directory structure created ({created_count} new directories)")
    return True

def verify_mt5_availability():
    """Check if MetaTrader 5 can be imported"""
    print("🔌 Checking MetaTrader 5 availability...")
    
    try:
        import MetaTrader5 as mt5
        if mt5.initialize():
            info = mt5.terminal_info()
            print("✅ MetaTrader 5 connection successful")
            print(f"   Terminal: {info.name}")
            print(f"   Version: {info.build}")
            mt5.shutdown()
            return True
        else:
            print("⚠️  MetaTrader 5 not connected - offline mode available")
            return False
    except ImportError:
        print("⚠️  MetaTrader 5 package not available - offline mode available")
        return False
    except Exception as e:
        print(f"⚠️  MetaTrader 5 check failed: {e}")
        return False

def verify_core_imports():
    """Verify that all core packages can be imported"""
    print("🔍 Verifying core package imports...")
    
    core_packages = [
        ('pandas', 'Data processing'),
        ('numpy', 'Numerical computations'), 
        ('matplotlib', 'Static visualizations'),
        ('seaborn', 'Statistical plots'),
        ('plotly', 'Interactive charts'),
        ('openpyxl', 'Excel file generation')
    ]
    
    failed_imports = []
    
    for package_name, description in core_packages:
        try:
            __import__(package_name)
            print(f"   ✅ {package_name:<12} - {description}")
        except ImportError:
            print(f"   ❌ {package_name:<12} - Failed to import")
            failed_imports.append(package_name)
    
    if failed_imports:
        print(f"\n❌ Failed to import: {', '.join(failed_imports)}")
        return False
    
    print("✅ All core packages imported successfully")
    return True

def create_config_file():
    """Create basic configuration file"""
    print("⚙️ Creating configuration file...")
    
    config_content = '''# Gold Trading Analysis Configuration
# Generated automatically by install.py

[analysis]
symbol = GOLD
default_timeframes = M1,M5,M15,M30,H1,H4,D1  
analysis_period = 25
min_sample_size = 25
confidence_level = 0.95

[paths]
data_export_dir = blog_data_exports
charts_dir = blog_data_exports/charts
tools_dir = blog_data_exports/downloadable_tools

[mt5]
timeout = 60000
max_bars = 50000

[visualization]
figure_dpi = 300
chart_style = seaborn
color_palette = viridis

[risk]
default_position_size = 0.1
default_risk_per_trade = 0.01
max_daily_risk = 0.05
'''
    
    config_path = Path("config/analysis_config.ini")
    config_path.parent.mkdir(exist_ok=True)
    
    with open(config_path, 'w') as f:
        f.write(config_content)
    
    print(f"✅ Configuration created: {config_path}")
    return True

def run_basic_tests():
    """Run basic functionality tests"""
    print("🧪 Running basic functionality tests...")
    
    # Test 1: Data processing
    try:
        import pandas as pd
        import numpy as np
        
        # Create test data
        test_data = pd.DataFrame({
            'timestamp': pd.date_range('2025-01-01', periods=100, freq='1min'),
            'open': np.random.uniform(2000, 2100, 100),
            'high': np.random.uniform(2000, 2100, 100),
            'low': np.random.uniform(2000, 2100, 100),
            'close': np.random.uniform(2000, 2100, 100),
        })
        
        print("   ✅ Data processing test passed")
        
    except Exception as e:
        print(f"   ❌ Data processing test failed: {e}")
        return False
    
    # Test 2: Visualization
    try:
        import matplotlib.pyplot as plt
        
        # Create test chart
        fig, ax = plt.subplots(figsize=(8, 6))
        ax.plot([1, 2, 3, 4], [1, 4, 2, 3])
        ax.set_title("Installation Test Chart")
        
        # Save to temp directory
        test_path = Path("temp/test_chart.png")
        fig.savefig(test_path, dpi=300, bbox_inches='tight')
        plt.close(fig)
        
        if test_path.exists():
            print("   ✅ Chart generation test passed")
            os.remove(test_path)  # Clean up
        else:
            print("   ❌ Chart generation test failed")
            return False
            
    except Exception as e:
        print(f"   ❌ Visualization test failed: {e}")
        return False
    
    # Test 3: Excel generation  
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Installation Test"
        ws['B1'] = "Success"
        
        test_path = Path("temp/test_excel.xlsx")
        wb.save(test_path)
        
        if test_path.exists():
            print("   ✅ Excel generation test passed")
            os.remove(test_path)  # Clean up
        else:
            print("   ❌ Excel generation test failed")
            return False
            
    except Exception as e:
        print(f"   ❌ Excel generation test failed: {e}")
        return False
    
    print("✅ All basic tests passed")
    return True

def print_next_steps():
    """Print next steps for the user"""
    print("\n" + "=" * 60)
    print("🎉 Installation Complete!")
    print("=" * 60)
    print()
    print("📋 Next Steps:")
    print("1. Run data export:     python mt5_data_exporter.py")
    print("2. Generate charts:     python chart_generator.py") 
    print("3. Create tools:        python interactive_analysis_tools.py")
    print("4. View results in:     blog_data_exports/ folder")
    print()
    print("📚 Documentation:")
    print("- Installation Guide:   docs/INSTALLATION.md")
    print("- API Reference:       docs/API_REFERENCE.md")
    print("- FAQ:                 docs/FAQ.md")
    print()
    print("🆘 Need Help?")
    print("- GitHub Issues:       https://github.com/TheHaywire/gold-trading-statistical-analysis/issues")
    print("- Documentation:       https://github.com/TheHaywire/gold-trading-statistical-analysis/blob/master/docs/")
    print()

def main():
    """Main installation process"""
    print_header()
    
    # Check prerequisites
    if not check_python_version():
        sys.exit(1)
    
    check_operating_system()
    print()
    
    # Install and setup
    if not install_requirements():
        print("\n❌ Installation failed at dependency installation")
        sys.exit(1)
    
    if not create_directory_structure():
        print("\n❌ Installation failed at directory creation")
        sys.exit(1)
    
    if not create_config_file():
        print("\n❌ Installation failed at configuration creation")
        sys.exit(1)
    
    print()
    
    # Verification
    if not verify_core_imports():
        print("\n❌ Installation failed at package verification")
        sys.exit(1)
    
    verify_mt5_availability()  # Non-critical
    print()
    
    if not run_basic_tests():
        print("\n❌ Installation failed at functionality tests")
        sys.exit(1)
    
    # Success!
    print_next_steps()

if __name__ == "__main__":
    main()